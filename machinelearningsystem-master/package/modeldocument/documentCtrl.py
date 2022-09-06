import openpyxl
from openpyxl.comments import Comment

class DocumentCtrl:

    def __init__(self):
        pass

    def MakeModelUseDataDoc (self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = [
            "commondata_001", "commondata_002", "commondata_003", "commondata_004", "commondata_005"
            , "commondata_006", "commondata_007", "commondata_008", "commondata_009", "commondata_010"
            , "commondata_011", "commondata_012", "commondata_013", "commondata_014", "commondata_015"
            , "uniquefloat_001", "uniquefloat_002", "uniquefloat_003", "uniquefloat_004", "uniquefloat_005"
            , "uniquefloat_006", "uniquefloat_007", "uniquefloat_008", "uniquefloat_009", "uniquefloat_010"
            , "uniquefloat_011", "uniquefloat_012", "uniquefloat_013", "uniquefloat_014", "uniquefloat_015"
            , "uniquefloat_016", "uniquefloat_017", "uniquefloat_018", "uniquefloat_019", "uniquefloat_020"
            , "uniquefloat_021", "uniquefloat_022", "uniquefloat_023", "uniquefloat_024", "uniquefloat_025"
            , "uniquefloat_026", "uniquefloat_027", "uniquefloat_028", "uniquefloat_029", "uniquefloat_030"
            , "uniquefloat_031", "uniquefloat_032", "uniquefloat_033", "uniquefloat_034", "uniquefloat_035"
            , "uniquefloat_036", "uniquefloat_037", "uniquefloat_038", "uniquefloat_039", "uniquefloat_040"
            , "uniquefloat_041", "uniquefloat_042", "uniquefloat_043", "uniquefloat_044", "uniquefloat_045"
            , "uniquefloat_046", "uniquefloat_047", "uniquefloat_048", "uniquefloat_049", "uniquefloat_050"
            , "uniquefloat_051", "uniquefloat_052", "uniquefloat_053", "uniquefloat_054", "uniquefloat_055"
            , "uniquefloat_056", "uniquefloat_057", "uniquefloat_058", "uniquefloat_059", "uniquefloat_060"
            , "uniquefloat_061", "uniquefloat_062", "uniquefloat_063", "uniquefloat_064", "uniquefloat_065"
            , "uniquefloat_066", "uniquefloat_067", "uniquefloat_068", "uniquefloat_069", "uniquefloat_070"
            , "uniquefloat_071", "uniquefloat_072", "uniquefloat_073", "uniquefloat_074", "uniquefloat_075"
            , "uniquefloat_076", "uniquefloat_077", "uniquefloat_078", "uniquefloat_079", "uniquefloat_080"
            , "uniquefloat_081", "uniquefloat_082", "uniquefloat_083", "uniquefloat_084", "uniquefloat_085"
            , "uniquefloat_086", "uniquefloat_087", "uniquefloat_088", "uniquefloat_089", "uniquefloat_090"
            , "uniquefloat_091", "uniquefloat_092", "uniquefloat_093", "uniquefloat_094", "uniquefloat_095"
            , "uniquefloat_096", "uniquefloat_097", "uniquefloat_098", "uniquefloat_099", "uniquefloat_100"
            , "uniquefloat_101", "uniquefloat_102", "uniquefloat_103", "uniquefloat_104", "uniquefloat_105"
            , "uniquefloat_106", "uniquefloat_107", "uniquefloat_108", "uniquefloat_109", "uniquefloat_110"
            , "uniquefloat_111", "uniquefloat_112", "uniquefloat_113", "uniquefloat_114", "uniquefloat_115"
            , "uniquefloat_116", "uniquefloat_117", "uniquefloat_118", "uniquefloat_119", "uniquefloat_120"
            , "uniquefloat_121", "uniquefloat_122", "uniquefloat_123", "uniquefloat_124", "uniquefloat_125"
            , "uniquefloat_126", "uniquefloat_127", "uniquefloat_128", "uniquefloat_129", "uniquefloat_130"
            , "uniquefloat_131", "uniquefloat_132", "uniquefloat_133", "uniquefloat_134", "uniquefloat_135"
            , "uniquefloat_136", "uniquefloat_137", "uniquefloat_138", "uniquefloat_139", "uniquefloat_140"
            , "uniquefloat_141", "uniquefloat_142", "uniquefloat_143", "uniquefloat_144", "uniquefloat_145"
            , "uniquefloat_146", "uniquefloat_147", "uniquefloat_148", "uniquefloat_149", "uniquefloat_150"
            , "uniquefloat_151", "uniquefloat_152", "uniquefloat_153", "uniquefloat_154", "uniquefloat_155"
            , "uniquefloat_156", "uniquefloat_157", "uniquefloat_158", "uniquefloat_159", "uniquefloat_160"
            , "uniquefloat_161", "uniquefloat_162", "uniquefloat_163", "uniquefloat_164", "uniquefloat_165"
            , "uniquefloat_166", "uniquefloat_167", "uniquefloat_168", "uniquefloat_169", "uniquefloat_170"
            , "uniquefloat_171", "uniquefloat_172", "uniquefloat_173", "uniquefloat_174", "uniquefloat_175"
            , "uniquefloat_176", "uniquefloat_177", "uniquefloat_178", "uniquefloat_179", "uniquefloat_180"
            , "uniquefloat_181", "uniquefloat_182", "uniquefloat_183", "uniquefloat_184", "uniquefloat_185"
            , "uniquefloat_186", "uniquefloat_187", "uniquefloat_188", "uniquefloat_189", "uniquefloat_190"
            , "uniquefloat_191", "uniquefloat_192", "uniquefloat_193", "uniquefloat_194", "uniquefloat_195"
            , "uniquefloat_196", "uniquefloat_197", "uniquefloat_198", "uniquefloat_199", "uniquefloat_200"
            , "uniquejson_001"]

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(initFilePath)
        for key in dataMap.keys():
            columnNumber = 4
            tableauDataArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for tableauData in tableauDataArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["project"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["version"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
                ws.cell(row=8, column=columnNumber, value="{}".format(tableauData["product"]["description"]))
                ws.cell(row=9, column=columnNumber, value="{}".format(tableauData["project"]["description"]))
                ws.cell(row=10, column=columnNumber, value="{}".format(tableauData["step"]["description"]))
                ws.cell(row=11, column=columnNumber, value="{}".format(tableauData["version"]["description"]))
                ws.cell(row=12, column=columnNumber, value="{}".format(tableauData["dt"]["description"]))

                if "commentMemo" in tableauData["tableInfo"].keys() :
                    ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(4))].comment = Comment(text="{}".format(tableauData["tableInfo"]["commentMemo"]), author="Code", height=100, width=500)

                rowNumber = 13
                for dataColumn in dataColumnArr:
                    rowNumber = rowNumber + 1
                    if dataColumn != tableauData[dataColumn]["description"]:
                        description = tableauData[dataColumn]["description"] if "description" in tableauData[dataColumn].keys() else ""
                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))

                        memo = tableauData[dataColumn]["memo"] if "memo" in tableauData[dataColumn].keys() else ""
                        commentMemo = ""
                        checkfunc = " , ".join(tableauData[dataColumn]["checkfunc"]) if "checkfunc" in tableauData[dataColumn].keys() else ""

                        if memo != "" or commentMemo != "" or checkfunc != "" :
                            ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = Comment(text="備註: {}\n其他備註: {}\n檢查方式: {}".format(memo, commentMemo,checkfunc), author="Code",height=100, width=500)

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)

    def MakeModelUseDataOmitDoc (self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = [
            "commondata_001", "commondata_002", "commondata_003", "commondata_004", "commondata_005"
            , "commondata_006", "commondata_007", "commondata_008", "commondata_009", "commondata_010"
            , "commondata_011", "commondata_012", "commondata_013", "commondata_014", "commondata_015"
            , "uniquefloat_001", "uniquefloat_002", "uniquefloat_003", "uniquefloat_004", "uniquefloat_005"
            , "uniquefloat_006", "uniquefloat_007", "uniquefloat_008", "uniquefloat_009", "uniquefloat_010"
            , "uniquefloat_011", "uniquefloat_012", "uniquefloat_013", "uniquefloat_014", "uniquefloat_015"
            , "uniquefloat_016", "uniquefloat_017", "uniquefloat_018", "uniquefloat_019", "uniquefloat_020"
            , "uniquefloat_021", "uniquefloat_022", "uniquefloat_023", "uniquefloat_024", "uniquefloat_025"
            , "uniquefloat_026", "uniquefloat_027", "uniquefloat_028", "uniquefloat_029", "uniquefloat_030"
            , "uniquefloat_031", "uniquefloat_032", "uniquefloat_033", "uniquefloat_034", "uniquefloat_035"
            , "uniquefloat_036", "uniquefloat_037", "uniquefloat_038", "uniquefloat_039", "uniquefloat_040"
            , "uniquefloat_041", "uniquefloat_042", "uniquefloat_043", "uniquefloat_044", "uniquefloat_045"
            , "uniquefloat_046", "uniquefloat_047", "uniquefloat_048", "uniquefloat_049", "uniquefloat_050"
            , "uniquefloat_051", "uniquefloat_052", "uniquefloat_053", "uniquefloat_054", "uniquefloat_055"
            , "uniquefloat_056", "uniquefloat_057", "uniquefloat_058", "uniquefloat_059", "uniquefloat_060"
            , "uniquefloat_061", "uniquefloat_062", "uniquefloat_063", "uniquefloat_064", "uniquefloat_065"
            , "uniquefloat_066", "uniquefloat_067", "uniquefloat_068", "uniquefloat_069", "uniquefloat_070"
            , "uniquefloat_071", "uniquefloat_072", "uniquefloat_073", "uniquefloat_074", "uniquefloat_075"
            , "uniquefloat_076", "uniquefloat_077", "uniquefloat_078", "uniquefloat_079", "uniquefloat_080"
            , "uniquefloat_081", "uniquefloat_082", "uniquefloat_083", "uniquefloat_084", "uniquefloat_085"
            , "uniquefloat_086", "uniquefloat_087", "uniquefloat_088", "uniquefloat_089", "uniquefloat_090"
            , "uniquefloat_091", "uniquefloat_092", "uniquefloat_093", "uniquefloat_094", "uniquefloat_095"
            , "uniquefloat_096", "uniquefloat_097", "uniquefloat_098", "uniquefloat_099", "uniquefloat_100"
            , "uniquefloat_101", "uniquefloat_102", "uniquefloat_103", "uniquefloat_104", "uniquefloat_105"
            , "uniquefloat_106", "uniquefloat_107", "uniquefloat_108", "uniquefloat_109", "uniquefloat_110"
            , "uniquefloat_111", "uniquefloat_112", "uniquefloat_113", "uniquefloat_114", "uniquefloat_115"
            , "uniquefloat_116", "uniquefloat_117", "uniquefloat_118", "uniquefloat_119", "uniquefloat_120"
            , "uniquefloat_121", "uniquefloat_122", "uniquefloat_123", "uniquefloat_124", "uniquefloat_125"
            , "uniquefloat_126", "uniquefloat_127", "uniquefloat_128", "uniquefloat_129", "uniquefloat_130"
            , "uniquefloat_131", "uniquefloat_132", "uniquefloat_133", "uniquefloat_134", "uniquefloat_135"
            , "uniquefloat_136", "uniquefloat_137", "uniquefloat_138", "uniquefloat_139", "uniquefloat_140"
            , "uniquefloat_141", "uniquefloat_142", "uniquefloat_143", "uniquefloat_144", "uniquefloat_145"
            , "uniquefloat_146", "uniquefloat_147", "uniquefloat_148", "uniquefloat_149", "uniquefloat_150"
            , "uniquefloat_151", "uniquefloat_152", "uniquefloat_153", "uniquefloat_154", "uniquefloat_155"
            , "uniquefloat_156", "uniquefloat_157", "uniquefloat_158", "uniquefloat_159", "uniquefloat_160"
            , "uniquefloat_161", "uniquefloat_162", "uniquefloat_163", "uniquefloat_164", "uniquefloat_165"
            , "uniquefloat_166", "uniquefloat_167", "uniquefloat_168", "uniquefloat_169", "uniquefloat_170"
            , "uniquefloat_171", "uniquefloat_172", "uniquefloat_173", "uniquefloat_174", "uniquefloat_175"
            , "uniquefloat_176", "uniquefloat_177", "uniquefloat_178", "uniquefloat_179", "uniquefloat_180"
            , "uniquefloat_181", "uniquefloat_182", "uniquefloat_183", "uniquefloat_184", "uniquefloat_185"
            , "uniquefloat_186", "uniquefloat_187", "uniquefloat_188", "uniquefloat_189", "uniquefloat_190"
            , "uniquefloat_191", "uniquefloat_192", "uniquefloat_193", "uniquefloat_194", "uniquefloat_195"
            , "uniquefloat_196", "uniquefloat_197", "uniquefloat_198", "uniquefloat_199", "uniquefloat_200"
            , "uniquejson_001" ]

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        dataTypeMap = {
            "string": "文字格式"
            , "integer": "數字格式"
            , "real": "數字格式"
            , "datetime": "日期時間格式"
            , "string_to_date": "日期格式"
            , "string_to_real": "數字格式"
        }

        wb = openpyxl.load_workbook(initFilePath)
        for key in dataMap.keys():
            columnNumber = 2
            tableauDataArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for tableauData in tableauDataArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["project"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["version"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
                if "commentMemo" in tableauData["tableInfo"].keys() :
                    ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(4))].comment = Comment(text="{}".format(tableauData["tableInfo"]["commentMemo"]), author="Code", height=100, width=500)
                rowNumber = 7
                for dataColumn in dataColumnArr:
                    if dataColumn != tableauData[dataColumn]["description"]:
                        rowNumber = rowNumber + 1
                        description = tableauData[dataColumn]["description"] if "description" in tableauData[dataColumn].keys() else ""
                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))

                        mainMemo = "【{}】{} ".format(dataTypeMap[tableauData[dataColumn]["datatype"]],dataColumn)
                        memo = tableauData[dataColumn]["memo"] if "memo" in tableauData[dataColumn].keys() else ""
                        commentMemo = ""
                        checkfunc = " , ".join(tableauData[dataColumn]["checkfunc"]) if "checkfunc" in tableauData[dataColumn].keys() else ""

                        if memo != "" or commentMemo != "" or checkfunc != "":
                            ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = Comment(text="{}\n備註: {}\n其他備註: {}\n檢查方式: {}".format(mainMemo,memo, commentMemo, checkfunc), author="Code", height=100, width=500)

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)

    def MakeDataSQL(self, projectInfoMain , makeInfo , tableName , product, project ,step, version, dataTime):
        dataColumnArr = [
            "commondata_001", "commondata_002", "commondata_003", "commondata_004", "commondata_005"
            , "commondata_006", "commondata_007", "commondata_008", "commondata_009", "commondata_010"
            , "commondata_011", "commondata_012", "commondata_013", "commondata_014", "commondata_015"
            , "uniquefloat_001", "uniquefloat_002", "uniquefloat_003", "uniquefloat_004", "uniquefloat_005"
            , "uniquefloat_006", "uniquefloat_007", "uniquefloat_008", "uniquefloat_009", "uniquefloat_010"
            , "uniquefloat_011", "uniquefloat_012", "uniquefloat_013", "uniquefloat_014", "uniquefloat_015"
            , "uniquefloat_016", "uniquefloat_017", "uniquefloat_018", "uniquefloat_019", "uniquefloat_020"
            , "uniquefloat_021", "uniquefloat_022", "uniquefloat_023", "uniquefloat_024", "uniquefloat_025"
            , "uniquefloat_026", "uniquefloat_027", "uniquefloat_028", "uniquefloat_029", "uniquefloat_030"
            , "uniquefloat_031", "uniquefloat_032", "uniquefloat_033", "uniquefloat_034", "uniquefloat_035"
            , "uniquefloat_036", "uniquefloat_037", "uniquefloat_038", "uniquefloat_039", "uniquefloat_040"
            , "uniquefloat_041", "uniquefloat_042", "uniquefloat_043", "uniquefloat_044", "uniquefloat_045"
            , "uniquefloat_046", "uniquefloat_047", "uniquefloat_048", "uniquefloat_049", "uniquefloat_050"
            , "uniquefloat_051", "uniquefloat_052", "uniquefloat_053", "uniquefloat_054", "uniquefloat_055"
            , "uniquefloat_056", "uniquefloat_057", "uniquefloat_058", "uniquefloat_059", "uniquefloat_060"
            , "uniquefloat_061", "uniquefloat_062", "uniquefloat_063", "uniquefloat_064", "uniquefloat_065"
            , "uniquefloat_066", "uniquefloat_067", "uniquefloat_068", "uniquefloat_069", "uniquefloat_070"
            , "uniquefloat_071", "uniquefloat_072", "uniquefloat_073", "uniquefloat_074", "uniquefloat_075"
            , "uniquefloat_076", "uniquefloat_077", "uniquefloat_078", "uniquefloat_079", "uniquefloat_080"
            , "uniquefloat_081", "uniquefloat_082", "uniquefloat_083", "uniquefloat_084", "uniquefloat_085"
            , "uniquefloat_086", "uniquefloat_087", "uniquefloat_088", "uniquefloat_089", "uniquefloat_090"
            , "uniquefloat_091", "uniquefloat_092", "uniquefloat_093", "uniquefloat_094", "uniquefloat_095"
            , "uniquefloat_096", "uniquefloat_097", "uniquefloat_098", "uniquefloat_099", "uniquefloat_100"
            , "uniquefloat_101", "uniquefloat_102", "uniquefloat_103", "uniquefloat_104", "uniquefloat_105"
            , "uniquefloat_106", "uniquefloat_107", "uniquefloat_108", "uniquefloat_109", "uniquefloat_110"
            , "uniquefloat_111", "uniquefloat_112", "uniquefloat_113", "uniquefloat_114", "uniquefloat_115"
            , "uniquefloat_116", "uniquefloat_117", "uniquefloat_118", "uniquefloat_119", "uniquefloat_120"
            , "uniquefloat_121", "uniquefloat_122", "uniquefloat_123", "uniquefloat_124", "uniquefloat_125"
            , "uniquefloat_126", "uniquefloat_127", "uniquefloat_128", "uniquefloat_129", "uniquefloat_130"
            , "uniquefloat_131", "uniquefloat_132", "uniquefloat_133", "uniquefloat_134", "uniquefloat_135"
            , "uniquefloat_136", "uniquefloat_137", "uniquefloat_138", "uniquefloat_139", "uniquefloat_140"
            , "uniquefloat_141", "uniquefloat_142", "uniquefloat_143", "uniquefloat_144", "uniquefloat_145"
            , "uniquefloat_146", "uniquefloat_147", "uniquefloat_148", "uniquefloat_149", "uniquefloat_150"
            , "uniquefloat_151", "uniquefloat_152", "uniquefloat_153", "uniquefloat_154", "uniquefloat_155"
            , "uniquefloat_156", "uniquefloat_157", "uniquefloat_158", "uniquefloat_159", "uniquefloat_160"
            , "uniquefloat_161", "uniquefloat_162", "uniquefloat_163", "uniquefloat_164", "uniquefloat_165"
            , "uniquefloat_166", "uniquefloat_167", "uniquefloat_168", "uniquefloat_169", "uniquefloat_170"
            , "uniquefloat_171", "uniquefloat_172", "uniquefloat_173", "uniquefloat_174", "uniquefloat_175"
            , "uniquefloat_176", "uniquefloat_177", "uniquefloat_178", "uniquefloat_179", "uniquefloat_180"
            , "uniquefloat_181", "uniquefloat_182", "uniquefloat_183", "uniquefloat_184", "uniquefloat_185"
            , "uniquefloat_186", "uniquefloat_187", "uniquefloat_188", "uniquefloat_189", "uniquefloat_190"
            , "uniquefloat_191", "uniquefloat_192", "uniquefloat_193", "uniquefloat_194", "uniquefloat_195"
            , "uniquefloat_196", "uniquefloat_197", "uniquefloat_198", "uniquefloat_199", "uniquefloat_200"
            , "uniquejson_001"
        ]

        sqlInit = """
-- [:DataName]
SELECT 
    AA.product as product -- 產品名稱
    , AA.project as project -- 計畫名稱
    , AA.step as step -- 資料步驟 
    , AA.version as version -- 資料版本
    , AA.dt as dt -- 資料時間 [:ColumnSQL] 
FROM [:TableName] AA
WHERE 1 = 1 
    AND AA.product = '[:Product]'
    AND AA.project = '[:Project]'
    AND AA.step = '[:Step]'
    AND AA.version = '[:Version]'
    AND AA.dt = '[:DataTime]'
Limit 10
        """

        columnInfoMap = eval(f"projectInfoMain.get{step}Info_{project}_{version}({makeInfo})[1]")
        columnSQL = ""
        for dataColumn in dataColumnArr:
            if dataColumn != columnInfoMap[dataColumn]["description"]:
                columnSQL = columnSQL + "\n    , AA.{} as {} -- {}".format(dataColumn,dataColumn, columnInfoMap[dataColumn]["description"])

        sql = sqlInit.replace("[:ColumnSQL]", columnSQL)
        sql = sql.replace("[:DataName]", f"產品：{product} 計畫：{project}　步驟 ：{step}　版本：{version}　時間：{dataTime}")
        sql = sql.replace("[:TableName]", tableName)

        sql = sql.replace("[:Product]", product)
        sql = sql.replace("[:Project]", project)
        sql = sql.replace("[:Step]", step)
        sql = sql.replace("[:Version]", version)
        sql = sql.replace("[:DataTime]", dataTime)

        return sql