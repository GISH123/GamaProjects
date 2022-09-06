import openpyxl
from openpyxl.comments import Comment

class DocumentCtrl:

    def __init__(self):
        pass

    def MakeModelExtractDoc (self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = [
            "commondata_1", "commondata_2", "commondata_3", "commondata_4", "commondata_5"
            , "commondata_6", "commondata_7", "commondata_8", "commondata_9", "commondata_10"
            , "commondata_11", "commondata_12", "commondata_13", "commondata_14", "commondata_15"
            , "uniqueint_1", "uniqueint_2", "uniqueint_3", "uniqueint_4", "uniqueint_5"
            , "uniqueint_6", "uniqueint_7", "uniqueint_8", "uniqueint_9", "uniqueint_10"
            , "uniqueint_11", "uniqueint_12", "uniqueint_13", "uniqueint_14", "uniqueint_15"
            , "uniquestr_1", "uniquestr_2", "uniquestr_3", "uniquestr_4", "uniquestr_5"
            , "uniquestr_6", "uniquestr_7", "uniquestr_8", "uniquestr_9", "uniquestr_10"
            , "uniquestr_11", "uniquestr_12", "uniquestr_13", "uniquestr_14", "uniquestr_15"
            , "uniquestr_16", "uniquestr_17", "uniquestr_18", "uniquestr_19", "uniquestr_20"
            , "uniquedbl_1", "uniquedbl_2", "uniquedbl_3", "uniquedbl_4", "uniquedbl_5"
            , "uniquedbl_6", "uniquedbl_7", "uniquedbl_8", "uniquedbl_9", "uniquedbl_10"
            , "uniquedbl_11", "uniquedbl_12", "uniquedbl_13", "uniquedbl_14", "uniquedbl_15"
            , "uniquedbl_16", "uniquedbl_17", "uniquedbl_18", "uniquedbl_19", "uniquedbl_20"
            , "uniquetime_1", "uniquetime_2", "uniquetime_3"
            , "otherstr_1", "otherstr_2", "otherstr_3", "otherstr_4", "otherstr_5"
            , "otherstr_6", "otherstr_7", "otherstr_8", "otherstr_9", "otherstr_10"
            , "uniquearray_1", "uniquearray_2", "uniquejson_1"
        ]

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(initFilePath)
        for key in dataMap.keys():
            columnNumber = 4
            tableauDataArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for tableauData in tableauDataArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["tableNumber"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataName"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
                ws.cell(row=8, column=columnNumber, value="{}".format(tableauData["game"]["description"]))
                ws.cell(row=9, column=columnNumber, value="{}".format(tableauData["dt"]["description"]))
                ws.cell(row=10, column=columnNumber, value="{}".format(tableauData["world"]["description"]))
                ws.cell(row=11, column=columnNumber, value="{}".format(tableauData["tableInfo"]["tableNumber"]))
                rowNumber = 12
                for dataColumn in dataColumnArr:
                    rowNumber = rowNumber + 1
                    if dataColumn != tableauData[dataColumn]["description"]:
                        description = tableauData[dataColumn]["description"] if "description" in tableauData[dataColumn].keys() else ""
                        pgmemo = tableauData[dataColumn]["pgmemo"] if "pgmemo" in tableauData[dataColumn].keys() else ""
                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))
                        if "memo" in tableauData[dataColumn].keys() and tableauData[dataColumn]["memo"] != "":
                            memo = tableauData[dataColumn]["memo"]
                            ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = Comment(text="{}\n{}".format(memo, pgmemo), author="Code",height=100, width=300)

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)

    def MakeBuReportDoc (self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = [
            "commondata_1", "commondata_2", "commondata_3", "commondata_4", "commondata_5"
            , "commondata_6", "commondata_7", "commondata_8", "commondata_9", "commondata_10"
            , "commondata_11", "commondata_12", "commondata_13", "commondata_14", "commondata_15"
            , "uniqueint_1", "uniqueint_2", "uniqueint_3", "uniqueint_4", "uniqueint_5"
            , "uniqueint_6", "uniqueint_7", "uniqueint_8", "uniqueint_9", "uniqueint_10"
            , "uniqueint_11", "uniqueint_12", "uniqueint_13", "uniqueint_14", "uniqueint_15"
            , "uniquestr_1", "uniquestr_2", "uniquestr_3", "uniquestr_4", "uniquestr_5"
            , "uniquestr_6", "uniquestr_7", "uniquestr_8", "uniquestr_9", "uniquestr_10"
            , "uniquestr_11", "uniquestr_12", "uniquestr_13", "uniquestr_14", "uniquestr_15"
            , "uniquestr_16", "uniquestr_17", "uniquestr_18", "uniquestr_19", "uniquestr_20"
            , "uniquedbl_1", "uniquedbl_2", "uniquedbl_3", "uniquedbl_4", "uniquedbl_5"
            , "uniquedbl_6", "uniquedbl_7", "uniquedbl_8", "uniquedbl_9", "uniquedbl_10"
            , "uniquedbl_11", "uniquedbl_12", "uniquedbl_13", "uniquedbl_14", "uniquedbl_15"
            , "uniquedbl_16", "uniquedbl_17", "uniquedbl_18", "uniquedbl_19", "uniquedbl_20"
            , "uniquetime_1", "uniquetime_2", "uniquetime_3"
            , "otherstr_1", "otherstr_2", "otherstr_3", "otherstr_4", "otherstr_5"
            , "otherstr_6", "otherstr_7", "otherstr_8", "otherstr_9", "otherstr_10"
            , "uniquearray_1", "uniquearray_2", "uniquejson_1"
        ]

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        wb = openpyxl.load_workbook(initFilePath)
        for key in dataMap.keys():
            columnNumber = 4
            tableauDataArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for tableauData in tableauDataArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["tableNumber"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataName"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
                ws.cell(row=8, column=columnNumber, value="{}".format(tableauData["game"]["description"]))
                ws.cell(row=9, column=columnNumber, value="{}".format(tableauData["dt"]["description"]))
                ws.cell(row=10, column=columnNumber, value="{}".format(tableauData["world"]["description"]))
                ws.cell(row=11, column=columnNumber, value="{}".format(tableauData["tableInfo"]["tableNumber"]))
                rowNumber = 12
                for dataColumn in dataColumnArr:
                    rowNumber = rowNumber + 1
                    if dataColumn != tableauData[dataColumn]["description"]:
                        description = tableauData[dataColumn]["description"] if "description" in tableauData[dataColumn].keys() else ""
                        pgmemo = tableauData[dataColumn]["pgmemo"] if "pgmemo" in tableauData[dataColumn].keys() else ""
                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))
                        if "memo" in tableauData[dataColumn].keys() and tableauData[dataColumn]["memo"] != "":
                            memo = tableauData[dataColumn]["memo"]
                            ws["{}{}".format(englishStr[columnNumber - 1 % 26], str(rowNumber))].comment = Comment(text="{}\n{}".format(memo, pgmemo), author="Code",height=100, width=300)


        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)

    def MakeTableauDoc (self,dataMap,initFilePath,outFilePath) :
        dataColumnArr = [
            "dt", "world", "game"
            , "commondata_1", "commondata_2", "commondata_3", "commondata_4", "commondata_5"
            , "commondata_6", "commondata_7", "commondata_8", "commondata_9", "commondata_10"
            , "commondata_11", "commondata_12", "commondata_13", "commondata_14", "commondata_15"
            , "uniqueint_1", "uniqueint_2", "uniqueint_3", "uniqueint_4", "uniqueint_5"
            , "uniqueint_6", "uniqueint_7", "uniqueint_8", "uniqueint_9", "uniqueint_10"
            , "uniqueint_11", "uniqueint_12", "uniqueint_13", "uniqueint_14", "uniqueint_15"
            , "uniquestr_1", "uniquestr_2", "uniquestr_3", "uniquestr_4", "uniquestr_5"
            , "uniquestr_6", "uniquestr_7", "uniquestr_8", "uniquestr_9", "uniquestr_10"
            , "uniquestr_11", "uniquestr_12", "uniquestr_13", "uniquestr_14", "uniquestr_15"
            , "uniquestr_16", "uniquestr_17", "uniquestr_18", "uniquestr_19", "uniquestr_20"
            , "uniquedbl_1", "uniquedbl_2", "uniquedbl_3", "uniquedbl_4", "uniquedbl_5"
            , "uniquedbl_6", "uniquedbl_7", "uniquedbl_8", "uniquedbl_9", "uniquedbl_10"
            , "uniquedbl_11", "uniquedbl_12", "uniquedbl_13", "uniquedbl_14", "uniquedbl_15"
            , "uniquedbl_16", "uniquedbl_17", "uniquedbl_18", "uniquedbl_19", "uniquedbl_20"
            , "uniquetime_1", "uniquetime_2", "uniquetime_3"
            , "otherstr_1", "otherstr_2", "otherstr_3", "otherstr_4", "otherstr_5"
            , "otherstr_6", "otherstr_7", "otherstr_8", "otherstr_9", "otherstr_10"
        ]

        englishStr = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        dataTypeMap = {
            "string": "文字格式"
            , "integer": "數字格式"
            , "real": "數字格式"
            , "datetime": "日期時間格式"
            , "string_to_date": "日期格式"
            , "string_to_real": "數字格式"
        }

        wb = openpyxl.load_workbook(initFilePath)
        for key in dataMap.keys():
            columnNumber = 2
            tableauDataArr = dataMap[key]
            ws_init = wb["Init"]
            ws = wb.copy_worksheet(ws_init)
            ws.title = key
            for tableauData in tableauDataArr:
                columnNumber = columnNumber + 1
                ws.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataCode"]))
                ws.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataName"]))
                ws.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
                ws.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
                ws.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
                rowNumber = 7
                for dataColumn in dataColumnArr:
                    if tableauData[dataColumn]["hidden"] == "false":
                        rowNumber = rowNumber + 1
                        description = tableauData[dataColumn]["description"] if "description" in tableauData[dataColumn].keys() else ""
                        ws.cell(row=rowNumber, column=columnNumber, value="{}".format(description))
                        if "memo" in tableauData[dataColumn].keys() and tableauData[dataColumn]["memo"] != "":
                            memo = tableauData[dataColumn]["memo"] if "memo" in tableauData[dataColumn].keys() else ""
                            commentMemo = "【{}】{} \n {}".format(dataTypeMap[tableauData[dataColumn]["datatype"]],dataColumn, memo)
                            ws["{}{}".format(englishStr[columnNumber - 1], str(rowNumber))].comment = Comment(text=commentMemo, author="Code",height=100, width=300)

        ws_init = wb["Init"]
        wb.remove(ws_init)
        wb.save(outFilePath)

    def MakeDataSQL(self ,tableInfoMain , makeInfo , tableName ,gameName , tableNumber, dataTime) :
        dataColumnArr = [
            "commondata_1", "commondata_2", "commondata_3", "commondata_4", "commondata_5"
            , "commondata_6", "commondata_7", "commondata_8", "commondata_9", "commondata_10"
            , "commondata_11", "commondata_12", "commondata_13", "commondata_14", "commondata_15"
            , "uniqueint_1", "uniqueint_2", "uniqueint_3", "uniqueint_4", "uniqueint_5"
            , "uniqueint_6", "uniqueint_7", "uniqueint_8", "uniqueint_9", "uniqueint_10"
            , "uniqueint_11", "uniqueint_12", "uniqueint_13", "uniqueint_14", "uniqueint_15"
            , "uniquestr_1", "uniquestr_2", "uniquestr_3", "uniquestr_4", "uniquestr_5"
            , "uniquestr_6", "uniquestr_7", "uniquestr_8", "uniquestr_9", "uniquestr_10"
            , "uniquestr_11", "uniquestr_12", "uniquestr_13", "uniquestr_14", "uniquestr_15"
            , "uniquestr_16", "uniquestr_17", "uniquestr_18", "uniquestr_19", "uniquestr_20"
            , "uniquedbl_1", "uniquedbl_2", "uniquedbl_3", "uniquedbl_4", "uniquedbl_5"
            , "uniquedbl_6", "uniquedbl_7", "uniquedbl_8", "uniquedbl_9", "uniquedbl_10"
            , "uniquedbl_11", "uniquedbl_12", "uniquedbl_13", "uniquedbl_14", "uniquedbl_15"
            , "uniquedbl_16", "uniquedbl_17", "uniquedbl_18", "uniquedbl_19", "uniquedbl_20"
            , "uniquetime_1", "uniquetime_2", "uniquetime_3"
            , "otherstr_1", "otherstr_2", "otherstr_3", "otherstr_4", "otherstr_5"
            , "otherstr_6", "otherstr_7", "otherstr_8", "otherstr_9", "otherstr_10"
            , "uniquearray_1", "uniquearray_2", "uniquejson_1"
        ]

        sqlInit = """
SELECT -- [:DataName]
    AA.game -- 產品名稱
    , AA.dt -- 資料時間
    , AA.world -- 世界資料
    , AA.tablenumber -- 報表編號 [:ColumnSQL] 
FROM [:TableName] AA
WHERE 1 = 1 
    AND AA.game = '[:GameName]'
    AND AA.dt = '[:DT]'
    AND AA.tablenumber = '[:TableNumber]'
Limit 10
        """

        columnInfoMap = eval(f"tableInfoMain.getBUReport{tableNumber}Info({makeInfo})[1]")
        columnSQL = ""
        for dataColumn in dataColumnArr:
            if dataColumn != columnInfoMap[dataColumn]["description"]:
                columnSQL = columnSQL + "\n    , AA.{} -- {}".format(dataColumn, columnInfoMap[dataColumn]["description"])

        sql = sqlInit.replace("[:ColumnSQL]", columnSQL)
        sql = sql.replace("[:GameName]", gameName)
        sql = sql.replace("[:TableNumber]", tableNumber)
        sql = sql.replace("[:DataName]", columnInfoMap["tableInfo"]["dataName"])
        sql = sql.replace("[:DT]", dataTime)
        sql = sql.replace("[:TableName]", tableName)

        return sql