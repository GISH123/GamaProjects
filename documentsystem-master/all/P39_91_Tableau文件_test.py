from info.maple.tableauinfo.TableauInfoMain import TableauInfoMain as TableauInfoMain
import time
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from openpyxl.comments import Comment

# 撈取相關
tableauInfoMain = TableauInfoMain()
tableauInfoMap = tableauInfoMain.getInitInfoMap()


makeInfo = {
    "tableauReportName" : "99_教學範本"
    , "gameName" :"maple"
    , "gameCHName" : "楓之谷"
    , "schemaName" : "maple_v"
    , "serverName": "10.10.99.151"
    , "serverPort": "5432"
    , "dbName": "bureport"
    , "userName": ""
}

dataMap = {
    "BU1000_登入與IP" : [
        tableauInfoMain.getBUReport1001Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1002Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1003Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1003Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1101Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1102Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1103Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1131Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1132Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1133Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1134Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1135Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1136Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1137Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1802Info(makeInfo)[1]
        , tableauInfoMain.getBUReport1804Info(makeInfo)[1]
    ]
    ,"BU6000_金流" : [
        tableauInfoMain.getBUReport6001Info(makeInfo)[1]
        , tableauInfoMain.getBUReport6002Info(makeInfo)[1]
        , tableauInfoMain.getBUReport6011Info(makeInfo)[1]
        , tableauInfoMain.getBUReport6012Info(makeInfo)[1]
        , tableauInfoMain.getBUReport6019Info(makeInfo)[1]
    ]
}

dataColumnArr = [
    "dt","world","game"
    ,"commondata_1","commondata_2","commondata_3","commondata_4","commondata_5"
    ,"commondata_6","commondata_7","commondata_8","commondata_9","commondata_10"
    ,"commondata_11","commondata_12","commondata_13","commondata_14","commondata_15"
    ,"uniqueint_1","uniqueint_2","uniqueint_3","uniqueint_4","uniqueint_5"
    ,"uniqueint_6","uniqueint_7","uniqueint_8","uniqueint_9","uniqueint_10"
    ,"uniqueint_11","uniqueint_12","uniqueint_13","uniqueint_14","uniqueint_15"
    ,"uniquestr_1","uniquestr_2","uniquestr_3","uniquestr_4","uniquestr_5"
    ,"uniquestr_6","uniquestr_7","uniquestr_8","uniquestr_9","uniquestr_10"
    ,"uniquestr_11","uniquestr_12","uniquestr_13","uniquestr_14","uniquestr_15"
    ,"uniquestr_16","uniquestr_17","uniquestr_18","uniquestr_19","uniquestr_20"
    ,"uniquedbl_1","uniquedbl_2","uniquedbl_3","uniquedbl_4","uniquedbl_5"
    ,"uniquedbl_6","uniquedbl_7","uniquedbl_8","uniquedbl_9","uniquedbl_10"
    ,"uniquedbl_11","uniquedbl_12","uniquedbl_13","uniquedbl_14","uniquedbl_15"
    ,"uniquedbl_16","uniquedbl_17","uniquedbl_18","uniquedbl_19","uniquedbl_20"
    ,"uniquetime_1","uniquetime_2","uniquetime_3"
    ,"otherstr_1","otherstr_2","otherstr_3","otherstr_4","otherstr_5"
    ,"otherstr_6","otherstr_7","otherstr_8","otherstr_9","otherstr_10"
]

englishStr="ABCDEFGHIJKLMNOPQRSTUVWXYZ"

dataTypeMap = {
    "string": "文字格式"
    , "integer": "數字格式"
    , "real": "數字格式"
    , "datetime": "日期時間格式"
    , "string_to_date": "日期格式"
    , "string_to_real": "數字格式"
}

wb = openpyxl.load_workbook('file/common/TableauDocInit.xlsx')
for key in dataMap.keys() :
    columnNumber = 2
    tableauDataArr = dataMap[key]
    ws_login = wb[key]
    for tableauData in tableauDataArr :
        columnNumber = columnNumber + 1
        print(tableauData["tableInfo"])
        ws_login.cell(row=2, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataCode"]))
        ws_login.cell(row=3, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataName"]))
        ws_login.cell(row=4, column=columnNumber, value="{}".format(tableauData["tableInfo"]["memo"]))
        ws_login.cell(row=5, column=columnNumber, value="{}".format(tableauData["tableInfo"]["startDate"] if "startDate" in tableauData["tableInfo"].keys() else "" ))
        ws_login.cell(row=6, column=columnNumber, value="{}".format(tableauData["tableInfo"]["endDate"] if "endDate" in tableauData["tableInfo"].keys() else ""))
        ws_login.cell(row=7, column=columnNumber, value="{}".format(tableauData["tableInfo"]["dataType"]))
        rowNumber = 7
        for dataColumn in dataColumnArr :
            if tableauData[dataColumn]["hidden"] == "false" :
                rowNumber = rowNumber + 1
                ws_login.cell(row=rowNumber, column=columnNumber, value="{}".format(tableauData[dataColumn]["description"]))
                memo = "【{}】\n {}".format(dataTypeMap[tableauData[dataColumn]["datatype"]],tableauData[dataColumn]["memo"])
                ws_login["{}{}".format(englishStr[columnNumber-1],str(rowNumber))].comment = Comment(text=memo,author="Code")

wb.save('file/maple/TableauDoc.xlsx')



