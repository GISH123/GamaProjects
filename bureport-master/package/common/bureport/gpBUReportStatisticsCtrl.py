import os , sys
from package.common.database.postgreCtrl import PostgresCtrl
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl


class GPBUReportStatisticsCtrl:

    def __init__(self):
        pass

    def MakeBUDetails(self, dataInfo):
        reportNameArr = dataInfo["reportNameArr"] if "reportNameArr" in dataInfo.keys() else []
        for reportName in reportNameArr:
            if reportName == "loginbasicreport" or reportName == "allreport":
                dataInfo["reportCodeArr"] = ["001", "101", "102", "103", "106", "107", "108","111"]
                dataInfo["xlsxFileOriName"] = "report_login_basic_ori.xlsx"
                dataInfo["xlsxFileSaveName"] = "report_login_basic_{}_{}.xlsx"
                dataInfo["xlsxSheetArr"] = ["Login", "Login_TW", "Login_HK"]
                self.MakeCommonReport(dataInfo)
            if reportName == "logintimereport" or reportName == "allreport":
                dataInfo["reportCodeArr"] = ["103", "111", "112", "113", "114"]
                dataInfo["xlsxFileOriName"] = "report_login_time_ori.xlsx"
                dataInfo["xlsxFileSaveName"] = "report_login_time_{}_{}.xlsx"
                dataInfo["xlsxSheetArr"] = ["Login", "Login_TW", "Login_HK"]
                self.MakeCommonReport(dataInfo)
            if reportName == "incomereport" or reportName == "allreport":
                dataInfo["reportCodeArr"] = ["001", "601"]
                dataInfo["xlsxFileOriName"] = "report_income_ori.xlsx"
                dataInfo["xlsxFileSaveName"] = "report_income_{}_{}.xlsx"
                dataInfo["xlsxSheetArr"] = ["Login_TW"]
                self.MakeCommonReport(dataInfo)
            if reportName == "incomelogintypereport" or reportName == "allreport":
                dataInfo["reportCodeArr"] = ["001", "602"]
                dataInfo["xlsxFileOriName"] = "report_income_logintype_ori.xlsx"
                dataInfo["xlsxFileSaveName"] = "report_income_logintype_{}_{}.xlsx"
                dataInfo["xlsxSheetArr"] = ["Login_TW"]
                self.MakeCommonReport(dataInfo)
            if reportName == "alldatareport" or reportName == "allreport":
                dataInfo["reportCodeArr"] = ["SELECT DISTINCT reportcode FROM bureportstatistics"]
                dataInfo["xlsxFileOriName"] = "report_all_data_ori.xlsx"
                dataInfo["xlsxFileSaveName"] = "report_all_data_{}_{}.xlsx"
                dataInfo["xlsxSheetArr"] = []
                self.MakeCommonReport(dataInfo)

    def MakeCommonReport(self , databaseInfo):
        gameName = databaseInfo["gameName"]
        maxDataName = databaseInfo["maxDataName"] if "maxDataName" in databaseInfo.keys() else []
        minDataName = databaseInfo["minDataName"] if "minDataName" in databaseInfo.keys() else []
        reportCodeArr = databaseInfo["reportCodeArr"] if "reportCodeArr" in databaseInfo.keys() else []
        dataNameArr = databaseInfo["dataNameArr"] if "dataNameArr" in databaseInfo.keys() else []
        xlsxFileOriName = databaseInfo["xlsxFileOriName"] if "xlsxFileOriName" in databaseInfo.keys() else "report_all_data_ori.xlsx"
        xlsxFileSaveName = databaseInfo["xlsxFileSaveName"] if "xlsxFileSaveName" in databaseInfo.keys() else "report_all_data_{}_{}.xlsx"
        xlsxSheetArr = databaseInfo["xlsxSheetArr"] if "xlsxSheetArr" in databaseInfo.keys() else []

        greenplumCtrl = PostgresCtrl(
            host="10.10.99.151"
            , port=int(os.getenv("POSTGRES_PORT"))
            , user="gpadmin"
            , password="!QAZ2wsx"
            , database="bureport"
            , schema=gameName
        )

        reportCodesStr = ""
        for reportCode in reportCodeArr :
            reportCodesStr += "'{}'".format(reportCode) if reportCodesStr == "" else ",'{}'".format(reportCode)

        dataNamesStr = ""
        for dataName in dataNameArr:
            dataNamesStr += "'{}'".format(dataName) if dataNamesStr == "" else ",'{}'".format(dataName)

        sqlReplaceArr = [
            ['[:GameName]', gameName]
            , ['[:ReportCodesStr]', reportCodesStr]
            , ['[:DataNamesStr]', dataNamesStr]
        ]

        sqlStr = """
            SELECT  
                AA.reportName
                --, AA.startdate
                --, AA.enddate
                --, AA.periodType
                , AA.gamename
                , AA.reportcode
                , AA.datatype1
                , AA.datatype2
                , AA.datatype3
                , AA.datatype4
                , AA.datatype5
                , AA.value     
            FROM [:GameName].bureportstatistics AA
            WHERE 1 = 1 
                AND AA.reportName in ([:DataNamesStr])
                AND AA.reportcode in ([:ReportCodesStr])
                AND AA.gamename = '[:GameName]'
        """
        for sqlReplace in sqlReplaceArr:
            sqlStr = sqlStr.replace(sqlReplace[0], sqlReplace[1])

        df = greenplumCtrl.searchSQL(sqlStr)
        df["value"] = df["value"].astype('float')

        rows = dataframe_to_rows(df)

        wb = openpyxl.load_workbook('file/{}'.format(xlsxFileOriName))
        ws = wb["Data"]
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if c_idx == 1:
                    continue
                if r_idx <= 1:
                    ws.cell(row=r_idx, column=c_idx - 1, value=value)
                elif r_idx == 2:
                    pass
                else:
                    ws.cell(row=r_idx - 1, column=c_idx - 1, value=value)


        for reportName in xlsxSheetArr :
            ws_login = wb[reportName]
            column_index = 10
            for dataName in dataNameArr :
                column_index = column_index + 1
                ws_login.cell(row=1, column=column_index, value="{}".format(dataName))

        wb.save('file/{}/xlsx_BUResult/{}'.format(gameName, xlsxFileSaveName.format(minDataName,maxDataName)))

    def getInitReportNameArr (self) :
        return ["loginbasicreport","logintimereport","incomereport","incomelogintypereport"]